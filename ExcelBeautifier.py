import sys
import os
import glob
import csv
import shutil
from colorama import init, Fore, Back, Style
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 初始化colorama，支持Windows系统
init(autoreset=True)


# 颜色代码定义（结合colorama和ANSI转义码）
class Colors:
    # 使用ANSI转义码代替可能不支持的Style属性
    HEADER = Fore.MAGENTA + Style.BRIGHT
    OKBLUE = Fore.BLUE
    OKGREEN = Fore.GREEN
    WARNING = Fore.YELLOW
    FAIL = Fore.RED
    ENDC = Style.RESET_ALL
    BRIGHT = Style.BRIGHT
    # 使用ANSI转义码实现下划线效果
    UNDERLINE = '\033[4m'
    PURPLE = Fore.MAGENTA


def print_colored_art():
    """打印彩色ASCII艺术和作者信息"""
    # 定义ASCII艺术图案
    ascii_art = [
        " _____              _ ____                   _   _  __ _           ",
        "| ____|_  _____ ___| | __ )  ___  __ _ _   _| |_(_)/ _(_) ___ _ __  ",
        "|  _| \\ \\/ / __/ _ \\ |  _ \\ / _ \\/ _` | | | | __| | |_| |/ _ \\ '__| ",
        "| |___ >  < (_|  __/ | |_) |  __/ (_| | |_| | |_| |  _| |  __/ |    ",
        "|_____/_/\\_\\___\\___|_|____/ \\___|\\__,_|\\__,_|\\__|_|_| |_|\\___|_|    "
    ]

    # 定义作者信息
    author_info = [
        "Author: Bifish",
        "Github: https://github.com/Bifish0"
    ]

    # 计算最长行的长度
    max_length = max(len(line) for line in ascii_art + author_info)

    # 打印顶部装饰线
    print(Fore.CYAN + "=" * (max_length + 4))

    # 打印ASCII艺术，使用绿色
    for line in ascii_art:
        padded_line = line.ljust(max_length)
        print(Fore.GREEN + f"| {padded_line} |")

    # 打印分隔线
    print(Fore.CYAN + "|" + "-" * (max_length + 2) + "|")

    # 打印作者信息，使用黄色
    for info in author_info:
        padded_info = info.center(max_length)
        print(Fore.YELLOW + f"| {padded_info} |")

    # 打印底部装饰线
    print(Fore.CYAN + "=" * (max_length + 4))

    # 重置颜色
    print(Style.RESET_ALL)


def print_colored(text, color):
    """带颜色打印文本"""
    print(f"{color}{text}{Colors.ENDC}")


def print_header(text):
    """打印标题样式文本"""
    print("\n" + "=" * 50)
    print_colored(f"{text:^50}", Colors.HEADER)
    print("=" * 50 + "\n")


def get_valid_directory(prompt, allow_default=False, default_dir=None):
    """获取有效的目录路径"""
    while True:
        dir_path = input(prompt).strip()

        # 处理默认目录情况
        if allow_default and not dir_path:
            if default_dir and os.path.isdir(default_dir):
                return default_dir
            else:
                print_colored("默认目录不存在，请输入有效的目录路径", Colors.WARNING)
                continue

        # 检查目录是否存在
        if os.path.isdir(dir_path):
            return os.path.abspath(dir_path)
        else:
            print_colored(f"错误：目录 '{dir_path}' 不存在，请重新输入", Colors.FAIL)


def csv_to_excel(csv_file_path, output_dir):
    """将CSV文件转换为Excel文件"""
    try:
        # 获取文件名（不含扩展名）
        file_name = os.path.splitext(os.path.basename(csv_file_path))[0]
        excel_file_path = os.path.join(output_dir, f"{file_name}.xlsx")

        # 创建备份
        if os.path.exists(excel_file_path):
            shutil.copy2(excel_file_path, f"{excel_file_path}.bak")
            print_colored(f"已创建备份文件: {excel_file_path}.bak", Colors.WARNING)

        # 创建一个新的Excel工作簿
        wb = Workbook()
        ws = wb.active

        # 读取CSV文件并写入Excel
        with open(csv_file_path, 'r', encoding='utf-8', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 保存Excel文件
        wb.save(excel_file_path)
        print_colored(f"已将CSV文件转换为Excel: {excel_file_path}", Colors.OKGREEN)
        return excel_file_path

    except Exception as e:
        print_colored(f"转换CSV文件 {csv_file_path} 时出错: {str(e)}", Colors.FAIL)
        return None


def beautify_excel(file_path, output_dir):
    """美化Excel文件的函数"""
    try:
        # 加载工作簿
        wb = load_workbook(file_path)

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        normal_font = Font(size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")

        # 处理每个工作表
        for sheet in wb.worksheets:
            # 如果工作表有数据
            if sheet.max_row > 0:
                # 设置标题行样式（第一行）
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # 调整列宽
            for col in range(1, sheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col)

                # 检查每一行的内容长度
                for row in range(1, sheet.max_row + 1):
                    cell = sheet[f"{column_letter}{row}"]
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                # 设置列宽（加一点缓冲）
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

            # 设置数据单元格样式
            for row in range(2, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = thin_border

                    # 尝试判断单元格内容类型设置对齐方式
                    if cell.value is not None:
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = center_alignment
                        else:
                            cell.alignment = left_alignment

        # 确定输出文件路径（覆盖原文件）
        output_file_path = os.path.join(output_dir, os.path.basename(file_path))

        # 创建备份
        if os.path.exists(output_file_path):
            shutil.copy2(output_file_path, f"{output_file_path}.bak")
            print_colored(f"已创建备份文件: {output_file_path}.bak", Colors.WARNING)

        # 保存美化后的文件
        wb.save(output_file_path)
        print_colored(f"已成功美化并保存至: {output_file_path}", Colors.OKGREEN)
        return True

    except Exception as e:
        print_colored(f"处理文件 {file_path} 时出错: {str(e)}", Colors.FAIL)
        return False


def select_files(file_list):
    """让用户通过序号选择文件，支持多个选择用英文逗号分隔，默认选择全部"""
    if not file_list:
        return []

    # 显示可选文件列表标题
    print_colored("\n可选文件列表：", Colors.OKGREEN + Colors.BRIGHT)
    # 显示序号和文件名
    for i, file_path in enumerate(file_list, 1):
        print_colored(f"{i}. {os.path.basename(file_path)}", Colors.PURPLE)

    # 获取用户选择
    while True:
        selection = input("\n请输入要处理的文件序号（多个用英文逗号分隔，直接回车选择全部）: ").strip()

        # 如果用户直接回车，默认选择全部文件
        if not selection:
            print_colored(f"已选择所有 {len(file_list)} 个文件进行处理", Colors.OKGREEN)
            return file_list

        try:
            # 解析选择的序号
            indices = [int(idx.strip()) - 1 for idx in selection.split(',')]
            # 验证序号有效性
            valid_indices = []
            for idx in indices:
                if 0 <= idx < len(file_list):
                    valid_indices.append(idx)
                else:
                    print_colored(f"警告：序号 {idx + 1} 无效，已忽略", Colors.WARNING)

            if not valid_indices:
                print_colored("没有有效的文件序号，请重新输入", Colors.WARNING)
                continue

            # 返回选中的文件
            selected_files = [file_list[idx] for idx in valid_indices]
            print_colored(f"\n已选择 {len(selected_files)} 个文件进行处理", Colors.OKGREEN)
            return selected_files

        except ValueError:
            print_colored("输入格式错误，请使用数字和英文逗号，如: 1,3,5", Colors.FAIL)


def process_files(source_dir, output_dir):
    """处理指定目录下的所有CSV和Excel文件"""
    # 获取所有CSV和Excel文件
    csv_files = glob.glob(os.path.join(source_dir, "*.csv"))
    excel_files = glob.glob(os.path.join(source_dir, "*.xlsx")) + glob.glob(os.path.join(source_dir, "*.xls"))

    all_files = csv_files + excel_files

    if not all_files:
        print_colored(f"在 {source_dir} 中没有找到Excel或CSV文件需要处理", Colors.WARNING)
        return

    # 让用户选择要处理的文件
    print_header("文件选择")
    print_colored(f"共发现 {len(csv_files)} 个CSV文件和 {len(excel_files)} 个Excel文件", Colors.OKBLUE)
    selected_files = select_files(all_files)

    if not selected_files:
        print_colored("未选择任何文件，处理终止", Colors.WARNING)
        return

    # 分离CSV和Excel文件
    selected_csv = [f for f in selected_files if f.lower().endswith('.csv')]
    selected_excel = [f for f in selected_files if f.lower().endswith(('.xlsx', '.xls'))]

    # 先处理CSV文件，转换为Excel
    if selected_csv:
        print_header("处理CSV文件")
        print_colored(f"开始处理 {len(selected_csv)} 个CSV文件...", Colors.OKBLUE)
        for csv_file in selected_csv:
            csv_to_excel(csv_file, output_dir)

    # 处理所有Excel文件
    if selected_excel:
        print_header("处理Excel文件")
        print_colored(f"开始美化 {len(selected_excel)} 个Excel文件...", Colors.OKBLUE)
        for file in selected_excel:
            beautify_excel(file, output_dir)

    print_header("处理完成")
    print_colored("所有选中的文件处理完毕", Colors.OKGREEN)


def check_and_install_libraries():
    """检查并安装所需库"""
    required_libraries = ['openpyxl', 'colorama']
    for lib in required_libraries:
        try:
            __import__(lib)
        except ImportError:
            print_colored(f"检测到未安装{lib}库，正在尝试安装...", Colors.WARNING)
            os.system(f"pip install {lib}")


if __name__ == "__main__":
    try:
        # 打印程序标题艺术
        print_colored_art()

        # 打印工具标题
        print_header("Excel/CSV 美化工具")

        # 检查并安装所需库
        check_and_install_libraries()

        # 获取源目录
        current_dir = os.getcwd()
        source_dir = get_valid_directory(
            f"{Colors.OKBLUE}请输入需要处理文件的目录: {Colors.ENDC}",
            allow_default=True,
            default_dir=current_dir
        )

        # 获取输出目录，默认使用源目录
        output_dir = get_valid_directory(
            f"{Colors.OKBLUE}请输入美化后文件的保存目录: {Colors.ENDC}",
            allow_default=True,
            default_dir=source_dir
        )

        # 处理文件
        process_files(source_dir, output_dir)
        input(f"\n{Colors.OKBLUE}按回车键退出...{Colors.ENDC}")

    except Exception as e:
        print_colored(f"发生错误: {e}", Colors.FAIL)
        sys.exit(1)
