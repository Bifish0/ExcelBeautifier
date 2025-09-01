import os
import glob
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def csv_to_excel(csv_file_path):
    """将CSV文件转换为Excel文件"""
    try:
        # 创建一个新的Excel工作簿
        wb = Workbook()
        ws = wb.active

        # 读取CSV文件并写入Excel
        with open(csv_file_path, 'r', encoding='utf-8', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 保存Excel文件，替换.csv为.xlsx
        excel_file_path = os.path.splitext(csv_file_path)[0] + '.xlsx'
        wb.save(excel_file_path)
        print(f"已将CSV文件转换为Excel: {excel_file_path}")
        return excel_file_path

    except Exception as e:
        print(f"转换CSV文件 {csv_file_path} 时出错: {str(e)}")
        return None


def beautify_excel(file_path):
    """美化Excel文件的函数，直接覆盖原文件"""
    try:
        # 加载工作簿
        wb = load_workbook(file_path)

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        normal_font = Font(size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")

        # 处理每个工作表
        for sheet in wb.worksheets:
            # 如果工作表有数据
            if sheet.max_row > 0:
                # 设置标题行样式（第一行）
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # 调整列宽
            for col in range(1, sheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col)

                # 检查每一行的内容长度
                for row in range(1, sheet.max_row + 1):
                    cell = sheet[f"{column_letter}{row}"]
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                # 设置列宽（加一点缓冲）
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

            # 设置数据单元格样式
            for row in range(2, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = thin_border

                    # 尝试判断单元格内容类型设置对齐方式
                    if cell.value is not None:
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = center_alignment
                        else:
                            cell.alignment = left_alignment

        # 直接覆盖原文件
        wb.save(file_path)
        print(f"已成功美化并覆盖文件: {file_path}")
        return True

    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        return False


def process_files():
    """处理当前目录下的所有CSV和Excel文件"""
    # 先处理CSV文件，转换为Excel
    csv_files = glob.glob("*.csv")

    if csv_files:
        print(f"找到 {len(csv_files)} 个CSV文件，开始转换...")
        for csv_file in csv_files:
            # 转换CSV为Excel
            excel_file = csv_to_excel(csv_file)

    # 处理所有Excel文件（包括原有的和从CSV转换来的）
    excel_files = glob.glob("*.xlsx") + glob.glob("*.xls")

    if not excel_files:
        print("没有找到Excel文件需要处理")
        return

    print(f"找到 {len(excel_files)} 个Excel文件，开始美化...")

    # 逐个处理Excel文件
    for file in excel_files:
        beautify_excel(file)

    print("所有文件处理完毕")


if __name__ == "__main__":
    # 检查并安装所需库
    required_libraries = ['openpyxl']
    for lib in required_libraries:
        try:
            __import__(lib)
        except ImportError:
            print(f"检测到未安装{lib}库，正在尝试安装...")
            os.system(f"pip install {lib}")

    process_files()
    input("按回车键退出...")
