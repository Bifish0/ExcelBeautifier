import os
import glob
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def get_valid_directory(prompt, allow_default=False, default_dir=None):
    """获取有效的目录路径"""
    while True:
        dir_path = input(prompt).strip()

        # 处理默认目录情况
        if allow_default and not dir_path:
            if default_dir and os.path.isdir(default_dir):
                return default_dir
            else:
                print("默认目录不存在，请输入有效的目录路径")
                continue

        # 检查目录是否存在
        if os.path.isdir(dir_path):
            return os.path.abspath(dir_path)
        else:
            print(f"错误：目录 '{dir_path}' 不存在，请重新输入")


def csv_to_excel(csv_file_path, output_dir):
    """将CSV文件转换为Excel文件"""
    try:
        # 获取文件名（不含扩展名）
        file_name = os.path.splitext(os.path.basename(csv_file_path))[0]
        excel_file_path = os.path.join(output_dir, f"{file_name}.xlsx")

        # 创建一个新的Excel工作簿
        wb = Workbook()
        ws = wb.active

        # 读取CSV文件并写入Excel
        with open(csv_file_path, 'r', encoding='utf-8', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 保存Excel文件
        wb.save(excel_file_path)
        print(f"已将CSV文件转换为Excel: {excel_file_path}")
        return excel_file_path

    except Exception as e:
        print(f"转换CSV文件 {csv_file_path} 时出错: {str(e)}")
        return None


def beautify_excel(file_path, output_dir):
    """美化Excel文件的函数"""
    try:
        # 加载工作簿
        wb = load_workbook(file_path)

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        normal_font = Font(size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")

        # 处理每个工作表
        for sheet in wb.worksheets:
            # 如果工作表有数据
            if sheet.max_row > 0:
                # 设置标题行样式（第一行）
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border

            # 调整列宽
            for col in range(1, sheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col)

                # 检查每一行的内容长度
                for row in range(1, sheet.max_row + 1):
                    cell = sheet[f"{column_letter}{row}"]
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                # 设置列宽（加一点缓冲）
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

            # 设置数据单元格样式
            for row in range(2, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.font = normal_font
                    cell.border = thin_border

                    # 尝试判断单元格内容类型设置对齐方式
                    if cell.value is not None:
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = center_alignment
                        else:
                            cell.alignment = left_alignment

        # 确定输出文件路径
        file_name = os.path.basename(file_path)
        output_file_path = os.path.join(output_dir, file_name)

        # 保存美化后的文件
        wb.save(output_file_path)
        print(f"已成功美化并保存至: {output_file_path}")
        return True

    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        return False


def process_files(source_dir, output_dir):
    """处理指定目录下的所有CSV和Excel文件"""
    # 先处理CSV文件，转换为Excel
    csv_files = glob.glob(os.path.join(source_dir, "*.csv"))

    if csv_files:
        print(f"在 {source_dir} 中找到 {len(csv_files)} 个CSV文件，开始转换...")
        for csv_file in csv_files:
            # 转换CSV为Excel并保存到输出目录
            csv_to_excel(csv_file, output_dir)

    # 处理所有Excel文件
    excel_files = glob.glob(os.path.join(source_dir, "*.xlsx")) + glob.glob(os.path.join(source_dir, "*.xls"))

    if not excel_files and not csv_files:
        print(f"在 {source_dir} 中没有找到Excel或CSV文件需要处理")
        return

    if excel_files:
        print(f"在 {source_dir} 中找到 {len(excel_files)} 个Excel文件，开始美化...")
        for file in excel_files:
            beautify_excel(file, output_dir)

    print("所有文件处理完毕")


if __name__ == "__main__":
    print("===== Excel/CSV 美化工具 =====")

    # 获取源目录
    current_dir = os.getcwd()
    source_dir = get_valid_directory(
        f"请输入需要处理文件的目录（默认：{current_dir}）: ",
        allow_default=True,
        default_dir=current_dir
    )

    # 获取输出目录，默认使用源目录
    output_dir = get_valid_directory(
        f"请输入美化后文件的保存目录（默认：{source_dir}）: ",
        allow_default=True,
        default_dir=source_dir
    )

    # 检查并安装所需库
    required_libraries = ['openpyxl']
    for lib in required_libraries:
        try:
            __import__(lib)
        except ImportError:
            print(f"检测到未安装{lib}库，正在尝试安装...")
            os.system(f"pip install {lib}")

    # 处理文件
    process_files(source_dir, output_dir)
    input("按回车键退出...")
